from glob import glob
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Font
from openpyxl.utils import get_column_letter
from time import sleep
from shutil import copyfile
import re, os, sys, openpyxl

emp_nums = {'조수안' : {"emp_num" : '1501011', "dept" : '바이오빅데이터', "team" : ''} \
            , '이덕형' : {"emp_num" : '1804232', "dept" : '바이오빅데이터', "team" : 'IT인프라개발'} \
            , '변하나' : {"emp_num" : '1906171', "dept" : '바이오빅데이터', "team" : 'IT인프라개발'} \
            , '신동민' : {"emp_num" : '2001131', "dept" : '바이오빅데이터', "team" : 'IT인프라개발'} \
            , '김수영' : {"emp_num" : '2010191', "dept" : '바이오빅데이터', "team" : 'IT인프라개발'} \
            , '한민영' : {"emp_num" : '2103152', "dept" : '바이오빅데이터', "team" : 'IT인프라개발'}}

argument = sys.argv
print(argument)

year = argument[1]
month = argument[2].rjust(2, '0')

path = 'D:\\document\\TimeSheet\\' + year + '\\' + month + '\\'
excels = glob(path + 'time sheet*.xlsx')

findex = 0
filename = ''
while True:
    findex += 1
    filename = 'team_v' + str(findex) + '.xlsx'
    if not os.path.exists(path + filename): break

filename = path + filename
first = True
for xlsx in excels:
    if "개인" in xlsx:
        copyfile(xlsx, filename)

wb_excel = openpyxl.load_workbook(filename)

for xlsx in excels:
    for k in emp_nums.keys():
        if k in xlsx:
            asis_sheet = wb_excel.active
            tobe_sheet = wb_excel.copy_worksheet(asis_sheet)
            tobe_sheet.title = k

            wb_temp = openpyxl.load_workbook(xlsx)
            ws_temp = wb_temp.worksheets[0]
            ws_temp

            for row in ws_temp:
                for cell in row:
                    tobe_sheet[cell.coordinate].value = cell.value

            break

sheets = wb_excel.sheetnames
ratio_col = 35
max_project_row = 4
for str_sht in sheets:
    if str_sht == '본부별':
        continue
    elif str_sht not in emp_nums:
        ts = wb_excel[str_sht]
        for i in range(10):
            t = i + 28
            tval = ts.cell(row = 4, column = t).value
            tval = str(tval)
            if bool(re.match('[가-힣]+', tval)):
                ratio_col = t
                break
        
        while True:
            max_project_row += 1
            tval = ts.cell(row = max_project_row, column = 1).value
            tval = str(tval)
            if not bool(re.match('[0-9]+', tval)): break
        continue

    sheet = wb_excel[str_sht]

    b2string = emp_nums[str_sht]["dept"]
    b2string += "/" + emp_nums[str_sht]["team"] if emp_nums[str_sht]["team"] else ''
    
    sheet['B2'] = b2string
    sheet['B2'].font = Font(bold=True, size=10, color="000000")
    sheet['B3'] = str_sht
    sheet['B3'].font = Font(bold=True, size=9, color="000000")

ss = wb_excel['본부별']
icol = 4
str_ratio_col = get_column_letter(ratio_col)
for k in emp_nums.keys():
    ss['B1'] = 'TIME SHEET_' + month.replace('0', '') + '월'
    irow = 1
    ss.cell(row = irow, column = icol).value = emp_nums[k]["emp_num"]
    irow += 1
    ss.cell(row = irow, column = icol).value = emp_nums[k]["dept"]
    irow += 1
    ss.cell(row = irow, column = icol).value = emp_nums[k]["team"]
    irow += 1
    ss.cell(row = irow, column = icol).value = k
    irow += 1
    for num in range(irow, max_project_row, 1):
        ss.cell(row = num, column = icol).value = '=' + k + '!' + str_ratio_col + str(num)
    icol += 1


wb_excel.remove(wb_excel[sheets[0]])
wb_excel.save(filename)
wb_excel.close()

print('finish --- ' + filename)