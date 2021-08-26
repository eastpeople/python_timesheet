from glob import glob
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, Font
from time import sleep
import re, os, openpyxl, xlwings as xw

emp_nums = {'조수안' : '1501011', '이덕형' : '1804232', '변하나' : '1906171' \
            , '신동민' : '2001131', '김수영' : '2010191', '한민영' : '2103152'}
excels = glob('time sheet*.xlsx')
wb_new = xw.Book()
first = True
for xlsx in excels:
    wb1 = xw.Book(xlsx)
    ws1 = wb1.sheets(1)
    print(xlsx + ' file working...')
    tnames = xlsx.split('_')
    emp_name = '이름없음'
    month_name = '0월'
    for name in tnames:
        if bool(re.match('[가-힣]+', name)):
            if len(name) > 2:
                emp_name = name
            else:
                month_name = name
    for sheet in wb_new.sheets:
        if emp_name in sheet.name: 
            sheet.delete()

    if first:
        wb1.sheets(2).api.Copy(Before=wb_new.sheets(1).api)
        first = False

    ws1.api.Copy(Before=wb_new.sheets(1).api)

    wb_new.sheets[0].name = emp_name
    wb1.close()
    
    

findex = 0
filename = ''
while True:
    findex += 1
    filename = 'team_v' + str(findex) + '.xlsx'
    if not os.path.exists(filename): break

wb_new.save(filename)
wb_new.close()
app = xw.App()
app.quit()

wb_excel = openpyxl.load_workbook(filename)
wb_excel.remove(wb_excel['Sheet1'])
sheets = wb_excel.sheetnames

for str_sht in sheets:
    if str_sht is '본부별':
        continue
    sheet = wb_excel[str_sht]
    sheet['B2'] = '바이오빅데이터/IT인프라개발'
    sheet['B2'].font = Font(bold=True, size=10, color="000000")
    sheet['B3'] = str_sht
    sheet['B3'].font = Font(bold=True, size=9, color="000000")

ss = wb_excel['본부별']
icol = 4
for k in emp_nums.keys():
    irow = 1
    ss.cell(row = irow, column = icol).value = emp_nums[k]
    irow += 1
    ss.cell(row = irow, column = icol).value = k
    icol += 1


wb_excel.save(filename)
wb_excel.close()

print('finish')